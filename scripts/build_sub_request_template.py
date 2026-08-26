"""
Build a locked-down SUB Request tracker template (.xlsx).

Every data column is constrained with Excel data validation so the sheet
rejects bad input at entry time, and the sheet itself is protected so only
the data-entry range can be edited.

Usage:  python scripts/build_sub_request_template.py [output.xlsx]
"""

import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------- settings --
OUTPUT = sys.argv[1] if len(sys.argv) > 1 else "SUB_Request_Tracker_Template.xlsx"

SHEET_DATA = "SUB Requests"
SHEET_LISTS = "Lists"
SHEET_RULES = "Rules"

FIRST_ROW = 2          # first data-entry row (also the sample row)
LAST_ROW = 1001        # last data-entry row covered by validation
PROTECT_PASSWORD = "Sub2026"

# Closed Date must fall inside July 2026.
DATE_MIN = date(2026, 7, 1)
DATE_MAX = date(2026, 7, 31)
EXCEL_EPOCH = date(1899, 12, 30)


def serial(d: date) -> int:
    """Excel 1900-system date serial number."""
    return (d - EXCEL_EPOCH).days


# ------------------------------------------------------------------ styles --
FONT = "Arial"
NAVY = "1F3864"
GREY_BG = "F2F2F2"
YELLOW = "FFF2CC"
RED_BG = "FFC7CE"
ORANGE_BG = "FFE699"

hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=NAVY)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

body_font = Font(name=FONT, size=10)
sample_font = Font(name=FONT, size=10, italic=True, color="808080")
title_font = Font(name=FONT, size=14, bold=True, color=NAVY)
bold_font = Font(name=FONT, size=10, bold=True)

thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

# ----------------------------------------------------------- column layout --
# key, header, width, number_format, sample value
COLUMNS = [
    ("sub_req",   "SUB Req Number",  16, "0",              1000123),
    ("mrc",       "MRC",             12, "#,##0.00",       249.50),
    ("party_id",  "Party ID",        14, "0",              8801234),
    ("closed",    "Closed Date",     14, "dd-mmm-yyyy",    date(2026, 7, 15)),
    ("kam_pt",    "KAM PT",          12, "@",              "PT-01"),
    ("kam_name",  "KAM Name",        22, "@",              "Sample KAM Name"),
    ("tl_name",   "TL Name",         22, "@",              "Sample TL Name"),
    ("dir_name",  "Director Name",   22, "@",              "Sample Director Name"),
    ("product",   "Product",         20, "@",              "Fiber Broadband"),
    ("rataplan",  "Rataplan Group",  20, "@",              "Group A"),
    ("remark",    "Remark",          40, "@",              "Sample remark - free text, optional"),
]
COL_IX = {key: i + 1 for i, (key, *_rest) in enumerate(COLUMNS)}
LAST_COL_LETTER = get_column_letter(len(COLUMNS))


def col(key: str) -> str:
    return get_column_letter(COL_IX[key])


def rng(key: str) -> str:
    c = col(key)
    return f"{c}{FIRST_ROW}:{c}{LAST_ROW}"


# ------------------------------------------------------ reference lists tab --
# Placeholder values - replace them with the real ones; the dropdowns and the
# named ranges below expand automatically as rows are added.
LISTS = [
    ("KAM PT", ["PT-01", "PT-02", "PT-03", "PT-04", "PT-05"]),
    ("KAM Name", ["Sample KAM Name", "KAM Two", "KAM Three"]),
    ("TL Name", ["Sample TL Name", "TL Two", "TL Three"]),
    ("Director Name", ["Sample Director Name", "Director Two"]),
    ("Product", ["Fiber Broadband", "Mobile Postpaid", "Fixed Voice",
                 "Business Internet", "Cloud / Managed Service"]),
    ("Rataplan Group", ["Group A", "Group B", "Group C", "Group D"]),
]
# defined-name per list column, in the same order as LISTS
LIST_NAMES = ["KAM_PT_List", "KAM_Name_List", "TL_Name_List",
              "Director_Name_List", "Product_List", "Rataplan_Group_List"]

wb = Workbook()

# =============================================================== Lists tab ==
ws_l = wb.active
ws_l.title = SHEET_LISTS
ws_l["A1"].value = None  # placeholder, headers written below

for i, (header, values) in enumerate(LISTS, start=1):
    letter = get_column_letter(i)
    cell = ws_l.cell(row=1, column=i, value=header)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = box
    for r, v in enumerate(values, start=2):
        c = ws_l.cell(row=r, column=i, value=v)
        c.font = body_font
        c.border = box
    ws_l.column_dimensions[letter].width = 24

ws_l.row_dimensions[1].height = 28
ws_l.freeze_panes = "A2"

note_row = max(len(v) for _, v in LISTS) + 3
ws_l.cell(row=note_row, column=1,
          value="Edit these lists to change what the dropdowns offer. "
                "Add new values directly under the last one - do NOT leave blank rows in "
                "the middle of a column, or the dropdown will stop at the gap.").font = Font(
    name=FONT, size=10, italic=True, color="C00000")

# Dynamic named ranges: Lists!$A$2 : INDEX(Lists!$A:$A, COUNTA(Lists!$A:$A))
for i, name in enumerate(LIST_NAMES, start=1):
    letter = get_column_letter(i)
    ref = (f"{SHEET_LISTS}!${letter}$2:"
           f"INDEX({SHEET_LISTS}!${letter}:${letter},COUNTA({SHEET_LISTS}!${letter}:${letter}))")
    wb.defined_names.add(DefinedName(name, attr_text=ref))

# ======================================================== Data-entry tab ====
ws = wb.create_sheet(SHEET_DATA, 0)

for i, (key, header, width, numfmt, sample) in enumerate(COLUMNS, start=1):
    letter = get_column_letter(i)
    c = ws.cell(row=1, column=i, value=header)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = hdr_align
    c.border = box
    ws.column_dimensions[letter].width = width

    # sample row
    s = ws.cell(row=FIRST_ROW, column=i, value=sample)
    s.font = sample_font
    s.number_format = numfmt
    s.border = box
    s.alignment = Alignment(vertical="center",
                            horizontal="left" if numfmt == "@" else "right",
                            wrap_text=(key == "remark"))

    # empty rows ready for input
    for r in range(FIRST_ROW + 1, LAST_ROW + 1):
        e = ws.cell(row=r, column=i)
        e.font = body_font
        e.number_format = numfmt
        e.border = box
        e.alignment = Alignment(vertical="center",
                                horizontal="left" if numfmt == "@" else "right",
                                wrap_text=(key == "remark"))

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{LAST_COL_LETTER}{LAST_ROW}"

# ------------------------------------------------------------ validations --
def add_dv(key, **kw):
    dv = DataValidation(allow_blank=True, showErrorMessage=True,
                        showInputMessage=True, errorStyle="stop", **kw)
    ws.add_data_validation(dv)
    dv.add(rng(key))
    return dv


sr = col("sub_req")
dv = add_dv(
    "sub_req",
    type="custom",
    # whole number, positive, and not already used elsewhere in the column
    formula1=(f"AND(ISNUMBER({sr}{FIRST_ROW}),{sr}{FIRST_ROW}=INT({sr}{FIRST_ROW}),"
              f"{sr}{FIRST_ROW}>0,"
              f"COUNTIF(${sr}${FIRST_ROW}:${sr}${LAST_ROW},{sr}{FIRST_ROW})=1)"),
)
dv.promptTitle = "SUB Req Number"
dv.prompt = "Whole number only. No text, no decimals, no duplicates."
dv.errorTitle = "Invalid SUB Req Number"
dv.error = ("This must be a positive whole number and must not already exist "
            "in this column. Text, decimals and duplicates are not allowed.")

dv = add_dv("mrc", type="decimal", operator="between",
            formula1="0", formula2="1000000")
dv.promptTitle = "MRC"
dv.prompt = "Numeric amount between 0 and 1,000,000. No text."
dv.errorTitle = "Invalid MRC"
dv.error = "MRC must be a number between 0 and 1,000,000. Text is not allowed."

dv = add_dv("party_id", type="whole", operator="between",
            formula1="1", formula2="999999999999")
dv.promptTitle = "Party ID"
dv.prompt = "Whole number only. No text, no spaces, no dashes."
dv.errorTitle = "Invalid Party ID"
dv.error = "Party ID must be a whole number. Text and symbols are not allowed."

dv = add_dv("closed", type="date", operator="between",
            formula1=str(serial(DATE_MIN)), formula2=str(serial(DATE_MAX)))
dv.promptTitle = "Closed Date"
dv.prompt = f"Date between {DATE_MIN:%d-%b-%Y} and {DATE_MAX:%d-%b-%Y} only."
dv.errorTitle = "Date outside July 2026"
dv.error = (f"Closed Date must be a real date between {DATE_MIN:%d-%b-%Y} and "
            f"{DATE_MAX:%d-%b-%Y}. Anything outside July 2026 is rejected.")

for key, name, label in [
    ("kam_pt", "KAM_PT_List", "KAM PT"),
    ("kam_name", "KAM_Name_List", "KAM Name"),
    ("tl_name", "TL_Name_List", "TL Name"),
    ("dir_name", "Director_Name_List", "Director Name"),
    ("product", "Product_List", "Product"),
    ("rataplan", "Rataplan_Group_List", "Rataplan Group"),
]:
    dv = add_dv(key, type="list", formula1=name)
    dv.promptTitle = label
    dv.prompt = f"Pick a value from the dropdown. Maintain the options on the '{SHEET_LISTS}' tab."
    dv.errorTitle = f"Invalid {label}"
    dv.error = (f"'{label}' only accepts values from the dropdown. "
                f"To add a new one, put it on the '{SHEET_LISTS}' tab first.")

dv = add_dv("remark", type="textLength", operator="lessThanOrEqual", formula1="250")
dv.promptTitle = "Remark"
dv.prompt = "Optional free text, up to 250 characters."
dv.errorTitle = "Remark too long"
dv.error = "Remark is limited to 250 characters."

# ------------------------------------------------- conditional formatting --
data_range = f"A{FIRST_ROW}:{LAST_COL_LETTER}{LAST_ROW}"
required_range = f"A{FIRST_ROW}:{col('rataplan')}{LAST_ROW}"

# missing mandatory value on a row that has been started
ws.conditional_formatting.add(
    required_range,
    FormulaRule(
        formula=[f'AND(COUNTA($A{FIRST_ROW}:${LAST_COL_LETTER}{FIRST_ROW})>0,'
                 f'A{FIRST_ROW}="")'],
        fill=PatternFill("solid", bgColor=RED_BG),
        stopIfTrue=False,
    ),
)
# banded look for readability
ws.conditional_formatting.add(
    data_range,
    FormulaRule(formula=[f"MOD(ROW(),2)=0"],
                fill=PatternFill("solid", bgColor=GREY_BG), stopIfTrue=False),
)

# ---------------------------------------------------------- sheet locking --
# Everything is locked by default; unlock only the data-entry grid.
unlocked = Protection(locked=False)
for r in range(FIRST_ROW, LAST_ROW + 1):
    for c_ix in range(1, len(COLUMNS) + 1):
        ws.cell(row=r, column=c_ix).protection = unlocked

ws.protection.set_password(PROTECT_PASSWORD)
ws.protection.sheet = True
ws.protection.selectLockedCells = False
ws.protection.selectUnlockedCells = False
ws.protection.formatCells = True
ws.protection.formatColumns = False
ws.protection.formatRows = False
ws.protection.insertRows = True
ws.protection.deleteRows = True
ws.protection.sort = False
ws.protection.autoFilter = False

# ================================================================ Rules tab =
ws_r = wb.create_sheet(SHEET_RULES)
ws_r.column_dimensions["A"].width = 22
ws_r.column_dimensions["B"].width = 18
ws_r.column_dimensions["C"].width = 78

ws_r["A1"] = "SUB Request Tracker - how to use this file"
ws_r["A1"].font = title_font
ws_r.merge_cells("A1:C1")

legend = [
    "",
    ("HOW TO FILL IT IN", "", ""),
    ("", "1.", f"Type only in the '{SHEET_DATA}' tab, rows {FIRST_ROW} to {LAST_ROW}. "
               "Everything else on that tab is locked."),
    ("", "2.", f"Row {FIRST_ROW} is a sample row (grey italic). Overwrite it with your first "
               "real record, or delete its contents."),
    ("", "3.", "A cell turns red when the row has been started but that value is still missing. "
               "Remark is the only optional column."),
    ("", "4.", f"To change what the dropdowns offer, edit the '{SHEET_LISTS}' tab - "
               "the dropdowns pick up new values automatically."),
    ("", "5.", f"To unlock the sheet (Review > Unprotect Sheet) the password is: {PROTECT_PASSWORD}"),
    "",
    ("VALIDATION RULES", "", ""),
]
row = 3
for item in legend:
    if item == "":
        row += 1
        continue
    a, b, c = item
    ws_r.cell(row=row, column=1, value=a).font = bold_font
    ws_r.cell(row=row, column=2, value=b).font = body_font
    cc = ws_r.cell(row=row, column=3, value=c)
    cc.font = body_font
    cc.alignment = Alignment(wrap_text=True, vertical="top")
    row += 1

rules = [
    ("SUB Req Number", "Whole number", "Positive whole number only. Text, decimals and "
                                       "duplicate numbers are rejected."),
    ("MRC", "Number", "Any number from 0 to 1,000,000 (decimals allowed). Text is rejected."),
    ("Party ID", "Whole number", "Whole number only. Text, spaces and dashes are rejected."),
    ("Closed Date", "Date", f"Must be a real date between {DATE_MIN:%d-%b-%Y} and "
                            f"{DATE_MAX:%d-%b-%Y}. Any date outside July 2026 is rejected."),
    ("KAM PT", "Dropdown", f"Must match a value on the '{SHEET_LISTS}' tab."),
    ("KAM Name", "Dropdown", f"Must match a value on the '{SHEET_LISTS}' tab."),
    ("TL Name", "Dropdown", f"Must match a value on the '{SHEET_LISTS}' tab."),
    ("Director Name", "Dropdown", f"Must match a value on the '{SHEET_LISTS}' tab."),
    ("Product", "Dropdown", f"Must match a value on the '{SHEET_LISTS}' tab."),
    ("Rataplan Group", "Dropdown", f"Must match a value on the '{SHEET_LISTS}' tab."),
    ("Remark", "Free text", "Optional. Up to 250 characters."),
]

hdr = ["Column", "Type", "Rule enforced by Excel"]
for i, h in enumerate(hdr, start=1):
    c = ws_r.cell(row=row, column=i, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = hdr_align
    c.border = box
ws_r.row_dimensions[row].height = 22
row += 1

for a, b, c in rules:
    ws_r.cell(row=row, column=1, value=a).font = bold_font
    ws_r.cell(row=row, column=2, value=b).font = body_font
    cc = ws_r.cell(row=row, column=3, value=c)
    cc.font = body_font
    cc.alignment = Alignment(wrap_text=True, vertical="top")
    for i in range(1, 4):
        ws_r.cell(row=row, column=i).border = box
    row += 1

row += 1
ws_r.cell(row=row, column=1, value="ASSUMPTIONS").font = bold_font
row += 1
assumptions = [
    "Dropdown values on the 'Lists' tab are placeholders - replace them with the real "
    "KAM / TL / Director / Product / Rataplan values.",
    "'KAM PT' is treated as a controlled code picked from a list. Tell me if it is a free "
    "employee number instead and I will switch it to a numeric rule.",
    f"'Closed Date' is locked to July 2026 as requested. Change DATE_MIN / DATE_MAX in "
    f"scripts/build_sub_request_template.py to move the window.",
    "Header spelling follows your request exactly, including 'Rataplan Group' "
    "(say the word if it should read 'Rate Plan Group').",
    f"Validation covers rows {FIRST_ROW}-{LAST_ROW}. Rows added past {LAST_ROW} are not validated.",
]
for a in assumptions:
    c = ws_r.cell(row=row, column=1, value="- " + a)
    c.font = body_font
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws_r.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws_r.row_dimensions[row].height = 28
    row += 1

for r in range(1, row + 1):
    for c_ix in range(1, 4):
        cell = ws_r.cell(row=r, column=c_ix)
        if cell.font.name is None:
            cell.font = body_font

ws_r.sheet_view.showGridLines = False

wb.save(OUTPUT)
print(f"wrote {OUTPUT}")
